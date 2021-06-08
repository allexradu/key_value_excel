import openpyxl as xl
import platform
import json

table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else 'excel/a.xlsx'
wb = xl.load_workbook(table_location)
sh = wb[wb.sheetnames[0]]

product_codes_duplicates = {}

for row in range(1, sh.max_row + 1):
    code = sh.cell(row, 2).value
    if code not in product_codes_duplicates.keys():
        product_codes_duplicates.update({code: False})
    else:
        product_codes_duplicates[code] = True

product_codes_parents = {}

for row in range(1, sh.max_row + 1):
    code = sh.cell(row, 2).value
    barcode = sh.cell(row, 1).value
    if product_codes_duplicates[code] is False:
        sh.cell(row, 3).value = barcode
    else:
        if code not in product_codes_parents.keys():
            sh.cell(row, 3).value = barcode
            product_codes_parents.update({code: barcode})
        else:
            sh.cell(row, 3).value = product_codes_parents[code]

wb.save(table_location)
