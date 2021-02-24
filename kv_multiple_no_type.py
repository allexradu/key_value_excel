import openpyxl as xl
import platform
import json

table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else 'excel/a.xlsx'
wb = xl.load_workbook(table_location)
sh = wb[wb.sheetnames[0]]
build_table_location = 'excel\\a_b.xlsx' if platform.system() == 'Windows' else 'excel/a_b.xlsx'
build_wb = xl.load_workbook(build_table_location)
build_sh = build_wb[build_wb.sheetnames[0]]
data = {}
col_dict = {}
current_new_col = 2
row_index = 2

for row in range(2, sh.max_row + 1):
    print('row: ', row)
    code = sh.cell(row, 1).value
    key = sh.cell(row, 2).value
    value = sh.cell(row, 3).value if sh.cell(row, 3).value is not None else ''

    if code not in data.keys():
        data.update({code: {'keys': [key], 'values': [value]}})
    else:
        data[code]['keys'].append(key)
        data[code]['values'].append(value)

with open('data.json', 'w') as fp:
    json.dump(data, fp, indent = 4)

for code in data.keys():
    build_sh.cell(row_index, 1).value = code
    print('build row', row_index)

    for a_type in data[code].keys():
        keys = data[code]['keys']
        values = data[code]['values']
        for i in range(len(keys)):
            if keys[i] not in col_dict.keys():
                col_dict.update({keys[i]: current_new_col})
                print(f'cell {col_dict[keys[i]]}-{row_index}: {values[i]} ')
                build_sh.cell(1, current_new_col).value = keys[i]
                build_sh.cell(row_index, current_new_col).value = values[i]
                current_new_col += 1
            else:
                build_sh.cell(row_index, col_dict[keys[i]]).value = values[i]
                print(f'cell {col_dict[keys[i]]}-{row_index}: {values[i]} ')

    row_index += 1

build_wb.save(build_table_location)
