import openpyxl as xl
import platform
import json
from datetime import datetime

table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else 'excel/a.xlsx'
table_location2 = 'excel\\a2.xlsx' if platform.system() == 'Windows' else 'excel/a2.xlsx'
wb = xl.load_workbook(table_location)
sh = wb[wb.sheetnames[0]]

cat1 = {}
cat2 = {}
cat3 = {}
cat4 = {}
parent = {}

for row in range(2, sh.max_row + 1):
    if sh.cell(row, 2).value == '0':
        cat1.update({sh.cell(row, 1).value: sh.cell(row, 3).value})

for row in range(2, sh.max_row + 1):
    if sh.cell(row, 2).value in cat1.keys():
        cat2.update({sh.cell(row, 1).value: sh.cell(row, 3).value})

for row in range(2, sh.max_row + 1):
    if sh.cell(row, 2).value in cat2.keys():
        cat3.update({sh.cell(row, 1).value: sh.cell(row, 3).value})

for row in range(2, sh.max_row + 1):
    if sh.cell(row, 2).value in cat3.keys():
        cat4.update({sh.cell(row, 1).value: sh.cell(row, 3).value})

for row in range(2, sh.max_row + 1):
    parent.update({sh.cell(row, 1).value: sh.cell(row, 2).value})

wb_build = xl.Workbook()
sh_build = wb_build[wb_build.sheetnames[0]]

for row in range(2, len(cat3.keys()) + 2):
    # cat4_code = list(cat4.keys())[row - 2]
    # sh_build.cell(row, 4).value = cat4[cat4_code]
    # cat3_code = parent[cat4_code]
    cat3_code = list(cat3.keys())[row - 2]
    sh_build.cell(row, 3).value = cat3[cat3_code]
    cat2_code = parent[cat3_code]
    sh_build.cell(row, 2).value = cat2[cat2_code]
    cat1_code = parent[cat2_code]
    sh_build.cell(row, 1).value = cat1[cat1_code]

wb_build.save(table_location2)