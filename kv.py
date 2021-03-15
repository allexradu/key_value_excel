import openpyxl as xl
import platform

table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else 'excel/a.xlsx'
wb = xl.load_workbook(table_location)
sh = wb[wb.sheetnames[0]]

keys = list(range(8, 10, 2))
values = list(range(9, 10, 2))
col_dict = {}
current_new_col = 10

for row in range(2, sh.max_row + 1):

    for col_no in range(len(keys)):
        print(f'row: {row}, col: {keys[col_no]}')
        key = sh.cell(row, keys[col_no]).value
        if key is not None:
            key = key.replace(':', '') if key.find(':') != -1 else key
            if key in col_dict.keys():
                col_val = col_dict[key]
                sh.cell(row, col_val).value = sh.cell(row, values[col_no]).value
            else:
                col_dict.update({key: current_new_col})
                sh.cell(1, current_new_col).value = key
                sh.cell(row, current_new_col).value = sh.cell(row, values[col_no]).value
                current_new_col += 1

wb.save(table_location)
