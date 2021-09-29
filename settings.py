import openpyxl as xl
import platform

table_location = 'excel\\set.xlsx' if platform.system() == 'Windows' else 'excel/set.xlsx'
wb = xl.load_workbook(table_location)
sh = wb[wb.sheetnames[0]]

f = open(r"excel\set.txt", "r", encoding = 'utf-8')
settings = f.read()

for row in range(2, sh.max_row + 1):
    if settings.find(sh.cell(row, 1).value) != -1:
        sh.cell(row,3).value = 'FOUND'
    else:
        sh.cell(row, 3).value = 'NOT FOUND'

wb.save(table_location)
