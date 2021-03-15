import openpyxl as xl
import platform
import json
import re

table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else 'excel/a.xlsx'
wb = xl.load_workbook(table_location)
sh = wb[wb.sheetnames[0]]
date1 = '07.02.2019 la 14:46'
date2 = '2021-02-17 10:32:12'
regex = r'(\d{2})\.(\d{2})\.(\d{4}) la (\d{1,2}):(\d{1,2})'
# v = re.match(regex, date1).groups()
# new_date = f'{v[2]}-{v[1]}-{v[0]} {v[3]}:{v[4]}:00'
# print(new_date)

for row in range(2, sh.max_row + 1):
    date = sh.cell(row, 2).value
    v = re.match(regex, date).groups()
    hours = v[3] if len(v[3]) == 2 else '0' + v[3]
    new_date = f'{v[2]}-{v[1]}-{v[0]} {hours}:{v[4]}:00'
    sh.cell(row, 3).value = new_date
wb.save(table_location)
