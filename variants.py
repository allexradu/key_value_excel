import openpyxl as xl
import platform
import json

table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else 'excel/a.xlsx'
wb = xl.load_workbook(table_location)
sh = wb[wb.sheetnames[0]]
build_table_location = 'excel\\a_b.xlsx' if platform.system() == 'Windows' else 'excel/a_b.xlsx'
build_wb = xl.load_workbook(build_table_location)
build_sh = build_wb[build_wb.sheetnames[0]]
data = {}
col_dict = {}
current_new_col = 2
row_index = 2
var_index = 2

for row in range(2, sh.max_row + 1):
    print('row: ', row)
    pid = sh.cell(row, 1).value
    code = sh.cell(row, 2).value
    p_name = sh.cell(row, 3).value
    color = sh.cell(row, 4).value
    image = sh.cell(row, 5).value

    if code not in data.keys():
        var_index = 2
        data.update({code: {'pid': pid, 'par_code': '', 'p_name': p_name, 'color': color, 'image': image}})
    else:
        data.update(
            {(code + '_' + str(var_index)): {'pid': str(pid), 'par_code': code, 'p_name': p_name, 'color': color,
                                             'image': image}})
        var_index += 1

with open('data.json', 'w') as fp:
    json.dump(data, fp, indent = 4)

for code in data.keys():
    build_sh.cell(row_index, 1).value = code
    build_sh.cell(row_index, 2).value = data[code]['par_code']
    build_sh.cell(row_index, 3).value = data[code]['pid']
    build_sh.cell(row_index, 4).value = data[code]['p_name']
    build_sh.cell(row_index, 5).value = data[code]['color']
    build_sh.cell(row_index, 6).value = data[code]['image']

    print('build row', row_index)
    row_index += 1

build_wb.save(build_table_location)
