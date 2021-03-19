import openpyxl as xl
import platform
import json

client_table_location = 'excel\\tier_client.xlsx' if platform.system() == 'Windows' else 'excel/tier_client.xlsx'
client_wb = xl.load_workbook(client_table_location)
client_sh = client_wb[client_wb.sheetnames[0]]

table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else 'excel/a.xlsx'
wb = xl.Workbook()
sh = wb[wb.sheetnames[0]]

client_table_product_id_column = 1
groups = range(1, 26)
rule_id_index = 1
build_row_index = 2

rule_id = 1
d_type = 2
prom_id = 3
product_ids = 4
supplier_id = 5
brands_ids = 6
category_ids = 7
group_ids = 8
company_id = 9
price_min = 10
price_max = 11
addition_percent = 12
addition_fixed = 13
final_price = 14
score = 15
public_site = 16
date_from = 17
date_to = 18
min_quantity = 19
max_quantity = 20
gift_product = 21
description = 22
transport = 23
gift_without_stock = 24
updated = 25
from_erp = 26

for group in groups:
    for column in range(2, client_sh.max_column + 1):
        for row in range(2, client_sh.max_row + 1):
            if client_sh.cell(row, column).value is not None:
                sh.cell(build_row_index, rule_id).value = rule_id_index
                sh.cell(build_row_index, d_type).value = 'discount'
                sh.cell(build_row_index, prom_id).value = 0
                sh.cell(build_row_index, product_ids).value = client_sh.cell(row, client_table_product_id_column).value
                print(client_sh.cell(row, client_table_product_id_column).value)
                sh.cell(build_row_index, supplier_id).value = 0
                sh.cell(build_row_index, brands_ids).value = ''
                sh.cell(build_row_index, category_ids).value = ''
                sh.cell(build_row_index, group_ids).value = group
                sh.cell(build_row_index, company_id).value = 0
                sh.cell(build_row_index, price_min).value = 0.00
                sh.cell(build_row_index, price_max).value = 0.00
                sh.cell(build_row_index, addition_percent).value = 0.00
                sh.cell(build_row_index, addition_fixed).value = 0.00
                sh.cell(build_row_index, final_price).value = client_sh.cell(row, column).value
                sh.cell(build_row_index, score).value = 30
                sh.cell(build_row_index, public_site).value = 0
                sh.cell(build_row_index, date_from).value = '2021-03-18 14:00:00'
                sh.cell(build_row_index, date_to).value = '2121-03-18 14:00:00'
                sh.cell(build_row_index, min_quantity).value = float(client_sh.cell(1, column).value)
                sh.cell(build_row_index, max_quantity).value = 0.00
                sh.cell(build_row_index, gift_product).value = ''
                sh.cell(build_row_index, description).value = ''
                sh.cell(build_row_index, transport).value = 0
                sh.cell(build_row_index, gift_without_stock).value = 0
                sh.cell(build_row_index, updated).value = 0
                sh.cell(build_row_index, from_erp).value = 0

                rule_id_index += 1
                build_row_index += 1

wb.save(table_location)
