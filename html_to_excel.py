import codecs

import requests
import lxml.html as lh
import openpyxl as xl
import platform
from lxml import etree
from bs4 import BeautifulSoup
import os

COLUMN_NO = 7
col_index = 1
row_index = 1

table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else 'excel/a.xlsx'
html_location = 'excel\\doc.html' if platform.system() == 'Windows' else 'excel/doc.htm'
wb = xl.load_workbook(table_location)
sh = wb[wb.sheetnames[0]]

file = codecs.open(html_location, "r", "utf-8").read()
# doc = lh.fromstring(file)
# tr_elements = doc.xpath('//tr')

soup = BeautifulSoup(file, "html.parser")
trs = soup.findChildren('td')
# print(trs[0].decode_contents())
# print(trs)

for el in range(len(trs)):
    sh.cell(row_index, col_index).value = trs[el].decode_contents()
    col_index += 1
    if col_index == COLUMN_NO:
        col_index = 1
        row_index += 1

wb.save(table_location)
