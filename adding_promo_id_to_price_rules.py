import openpyxl as xl
import platform
import json




table_location = 'excel\\a.xlsx' if platform.system() == 'Windows' else 'excel/a.xlsx'
wb = xl.load_workbook(table_location)
sh = wb[wb.sheetnames[0]]

ids = [
    '9997',
    '10002',
    '9986',
    '9988',
    '9922',
    '9831',
    '9887',
    '9990',
    '10138',
    '9892',
    '10278',
    '9998'
]

promo_id = 23

for row in range(2, sh.max_row + 1):
    id_cell = str(sh.cell(row, 4).value)
    if id_cell is not None:
        if id_cell.find(',') == -1:
            if id_cell in ids:
                sh.cell(row, 3).value = promo_id
        else:
            id_list = id_cell.split(',')
            for pid in id_list:
                if pid in ids:
                    sh.cell(row, 3).value = promo_id
                    break

wb.save(table_location)
