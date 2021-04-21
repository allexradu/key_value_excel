import openpyxl as xl
import platform
import json
from datetime import datetime

# table_location = 'excel\\hdz.xlsx' if platform.system() == 'Windows' else 'excel/hdz.xlsx'
# wb = xl.load_workbook(table_location)
# sh = wb[wb.sheetnames[0]]
#
# table_location2 = 'excel\\iduri.xlsx' if platform.system() == 'Windows' else 'excel/iduri.xlsx'
# wb_codes = xl.load_workbook(table_location2)
# sh_codes = wb[wb_codes.sheetnames[0]]
#
# ids = [sh_codes.cell(row, 1) for row in range(1, sh_codes.max_row + 1)]
#
# time = datetime.now()
#
DELETE = True
# CODES_COLUMN = 8
# DELETE_COLUMN = 12
# data = {}

# for row in range(2, sh.max_row + 1):
#     print(f'add to memory: {row} / {sh.max_row + 1}')
#     code = sh.cell(row, CODES_COLUMN).value
#     if code not in data.keys():
#         data.update({code: False})
#     else:
#         data.update({code: True})

# with open('data.json', 'w') as fp:
#     json.dump(data, fp, indent = 4)

with open('data.json') as json_file:
    data = json.load(json_file)

if DELETE is False:
    pass
# for row in range(2, sh.max_row + 1):
#     print(f'add delete: {row} / {sh.max_row + 1}')
#     code = sh.cell(row, CODES_COLUMN).value
#     if code is not None:
#         is_duplicate = data[code]
#         print(is_duplicate)
#
#         if is_duplicate:
#             sh.cell(row, DELETE_COLUMN).value = 'DELETE'
else:
    temp = 0
    for key in data.keys():
        if data[key] is True:
            temp += 1
    print(temp)

    # for row in range(2, sh.max_row + 1):
    #     print(f'checking delete: {row} / {sh.max_row + 1} time: {datetime.now() - time}')
    #     code = str(sh.cell(row, CODES_COLUMN).value)
    #     if code is not None:
    #         try:
    #             is_duplicate = data[code]
    #             print(is_duplicate)
    #             if is_duplicate and code not in ids:
    #                 print(f'DELETED: {row} / {sh.max_row + 1} time: {datetime.now() - time}')
    #                 sh.delete_rows(row)
    #         except:
    #             pass

# wb.save(table_location)
