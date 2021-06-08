import openpyxl as xl
import platform
import json
from datetime import datetime

table_location_edo = 'excel\\a_edo.xlsx' if platform.system() == 'Windows' else 'excel/a_edo.xlsx'
table_location_max = 'excel\\a_max.xlsx' if platform.system() == 'Windows' else 'excel/a_max.xlsx'
table_location2 = 'excel\\a2.xlsx' if platform.system() == 'Windows' else 'excel/a2.xlsx'
wb_edo = xl.load_workbook(table_location_edo)
sh_edo = wb_edo[wb_edo.sheetnames[0]]
wb_max = xl.load_workbook(table_location_max)
sh_max = wb_max[wb_max.sheetnames[0]]

cat1_edo = {}
cat2_edo = {}
cat3_edo = {}
cat4_edo = {}
parent_edo = {}
cat1_max = {}
cat2_max = {}
cat3_max = {}
cat4_max = {}
parent_max = {}
row_index = 2

for row in range(2, sh_edo.max_row + 1):
    if sh_edo.cell(row, 2).value == 0:
        cat1_edo.update({sh_edo.cell(row, 1).value: sh_edo.cell(row, 3).value})

for row in range(2, sh_edo.max_row + 1):
    if sh_edo.cell(row, 2).value in cat1_edo.keys():
        cat2_edo.update({sh_edo.cell(row, 1).value: sh_edo.cell(row, 3).value})

for row in range(2, sh_edo.max_row + 1):
    if sh_edo.cell(row, 2).value in cat2_edo.keys():
        cat3_edo.update({sh_edo.cell(row, 1).value: sh_edo.cell(row, 3).value})

for row in range(2, sh_edo.max_row + 1):
    if sh_edo.cell(row, 2).value in cat3_edo.keys():
        cat4_edo.update({sh_edo.cell(row, 1).value: sh_edo.cell(row, 3).value})

for row in range(2, sh_edo.max_row + 1):
    parent_edo.update({sh_edo.cell(row, 1).value: sh_edo.cell(row, 2).value})

for row in range(2, sh_max.max_row + 1):
    if sh_max.cell(row, 2).value == 0:
        cat1_max.update({sh_max.cell(row, 1).value: sh_max.cell(row, 3).value})

for row in range(2, sh_max.max_row + 1):
    if sh_max.cell(row, 2).value in cat1_max.keys():
        cat2_max.update({sh_max.cell(row, 1).value: sh_max.cell(row, 3).value})

for row in range(2, sh_max.max_row + 1):
    if sh_max.cell(row, 2).value in cat2_max.keys():
        cat3_max.update({sh_max.cell(row, 1).value: sh_max.cell(row, 3).value})

for row in range(2, sh_max.max_row + 1):
    if sh_max.cell(row, 2).value in cat3_max.keys():
        cat4_max.update({sh_max.cell(row, 1).value: sh_max.cell(row, 3).value})

for row in range(2, sh_max.max_row + 1):
    parent_max.update({sh_max.cell(row, 1).value: sh_max.cell(row, 2).value})

wb_build = xl.Workbook()
sh_build = wb_build[wb_build.sheetnames[0]]

print(cat1_edo)

for code in cat1_edo.keys():
    print(code)
    if cat1_edo[code] in cat1_max.values():
        category_name = cat1_edo[code]

        sh_build.cell(row_index, 1).value = code
        code_max_index = list(cat1_max.values()).index(category_name)

        sh_build.cell(row_index, 2).value = list(cat1_max.keys())[code_max_index]
        row_index += 1

for code in cat2_edo.keys():
    if cat2_edo[code] in cat2_max.values():
        category_name = cat2_edo[code]
        sh_build.cell(row_index, 1).value = code
        code_max_index = list(cat2_max.values()).index(category_name)
        sh_build.cell(row_index, 2).value = list(cat2_max.keys())[code_max_index]
        row_index += 1

for code in cat3_edo.keys():
    if cat3_edo[code] in cat3_max.values():
        category_name = cat3_edo[code]
        sh_build.cell(row_index, 1).value = code
        code_max_index = list(cat3_max.values()).index(category_name)
        sh_build.cell(row_index, 2).value = list(cat3_max.keys())[code_max_index]
        row_index += 1

for code in cat4_edo.keys():
    if cat4_edo[code] in cat4_max.values():
        category_name = cat4_edo[code]
        sh_build.cell(row_index, 1).value = code
        code_max_index = list(cat4_max.values()).index(category_name)
        sh_build.cell(row_index, 2).value = list(cat4_max.keys())[code_max_index]
        row_index += 1

print(cat1_edo)


# for row in range(2, len(cat3_edo.keys()) + 2):
#     cat4_code = list(cat4_edo.keys())[row - 2]
#     sh_build.cell(row, 8).value = cat4_code
#     sh_build.cell(row, 7).value = cat4_edo[cat4_code]
#     cat3_code = parent_edo[cat4_code]
#     # cat3_code = list(cat3.keys())[row - 2]
#     sh_build.cell(row, 6).value = cat3_code
#     sh_build.cell(row, 5).value = cat3_edo[cat3_code]
#     cat2_code = parent_edo[cat3_code]
#     sh_build.cell(row, 4).value = cat2_code
#     sh_build.cell(row, 3).value = cat2_edo[cat2_code]
#     cat1_code = parent_edo[cat2_code]
#     sh_build.cell(row, 2).value = cat1_code
#     sh_build.cell(row, 1).value = cat1_edo[cat1_code]

wb_build.save(table_location2)
